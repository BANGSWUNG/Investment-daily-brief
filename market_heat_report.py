#!/usr/bin/env python3
"""
시장 열기 트래커 (Market Heat Tracker)
─────────────────────────────────────
yfinance → ATR / RVOL / OBV / 거래대금 자동 계산 → Excel 보고서 생성
GitHub Actions + 이메일 자동 발송 지원

비용: $0 (yfinance 무료, API 키 불필요)
"""

import argparse
import json
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════
# 설정: 추적 대상 티커 (ETF 프록시)
# ══════════════════════════════════════════════════════════════

COUNTRY_ETFS = {
    "🇺🇸 미국 (S&P500)": "SPY",
    "🇺🇸 미국 (나스닥)": "QQQ",
    "🇰🇷 한국": "EWY",
    "🇯🇵 일본": "EWJ",
    "🇨🇳 중국": "FXI",
    "🇹🇼 대만": "EWT",
    "🇮🇳 인도": "INDA",
    "🇧🇷 브라질": "EWZ",
    "🇩🇪 독일": "EWG",
    "🇪🇺 유럽": "VGK",
}

SECTOR_ETFS = {
    "기술": "XLK",
    "금융": "XLF",
    "에너지": "XLE",
    "헬스케어": "XLV",
    "산업재": "XLI",
    "커뮤니케이션": "XLC",
    "경기소비재": "XLY",
    "필수소비재": "XLP",
    "소재": "XLB",
    "유틸리티": "XLU",
    "부동산": "XLRE",
}

CRYPTO_TICKERS = {
    "비트코인": "BTC-USD",
    "이더리움": "ETH-USD",
    "솔라나": "SOL-USD",
}

COMMODITY_ETFS = {
    "금": "GLD",
    "원유 (WTI)": "USO",
    "은": "SLV",
    "천연가스": "UNG",
}

ATR_PERIOD = 14
RVOL_BASELINE = 20  # 20일 평균 대비
LOOKBACK_DAYS = 60  # 계산용 과거 데이터


# ══════════════════════════════════════════════════════════════
# 데이터 수집 & 지표 계산
# ══════════════════════════════════════════════════════════════

def fetch_and_calc(tickers: dict, category: str, is_crypto: bool = False) -> pd.DataFrame:
    """티커 딕셔너리 → yfinance 다운로드 → 지표 계산 → DataFrame 반환"""
    end = datetime.now()
    start = end - timedelta(days=LOOKBACK_DAYS + 10)

    rows = []
    for name, ticker in tickers.items():
        try:
            df = yf.download(ticker, start=start, end=end, progress=False, auto_adjust=True)
            if df.empty or len(df) < ATR_PERIOD + 5:
                print(f"  ⚠ {ticker}: 데이터 부족, 건너뜀")
                continue

            # Flatten MultiIndex columns if present
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)

            close = df["Close"]
            high = df["High"]
            low = df["Low"]
            volume = df["Volume"]

            # --- ATR(14) ---
            tr = pd.concat([
                high - low,
                (high - close.shift(1)).abs(),
                (low - close.shift(1)).abs()
            ], axis=1).max(axis=1)
            atr = tr.rolling(ATR_PERIOD).mean()
            atr_latest = atr.iloc[-1]
            atr_pct = (atr_latest / close.iloc[-1]) * 100  # ATR%

            # --- RVOL ---
            vol_avg_20 = volume.rolling(RVOL_BASELINE).mean()
            rvol = volume.iloc[-1] / vol_avg_20.iloc[-1] if vol_avg_20.iloc[-1] > 0 else 0

            # --- OBV ---
            obv = (volume * ((close.diff() > 0).astype(int) * 2 - 1)).cumsum()
            obv_latest = obv.iloc[-1]
            # OBV 5일 변화율
            obv_5d_chg = ((obv.iloc[-1] - obv.iloc[-6]) / abs(obv.iloc[-6]) * 100
                          if len(obv) >= 6 and obv.iloc[-6] != 0 else 0)

            # --- 거래대금 ($ Volume) ---
            # 크립토: yfinance Volume이 이미 USD 거래량이므로 그대로 사용
            # ETF: price × volume = 달러 거래대금
            if is_crypto:
                dollar_vol_today = volume.iloc[-1]
                dollar_vol_avg = volume.rolling(RVOL_BASELINE).mean().iloc[-1]
            else:
                dollar_vol_today = close.iloc[-1] * volume.iloc[-1]
                dollar_vol_avg = (close * volume).rolling(RVOL_BASELINE).mean().iloc[-1]

            # --- 가격 변동 ---
            pct_1d = ((close.iloc[-1] / close.iloc[-2]) - 1) * 100 if len(close) >= 2 else 0
            pct_5d = ((close.iloc[-1] / close.iloc[-6]) - 1) * 100 if len(close) >= 6 else 0
            pct_20d = ((close.iloc[-1] / close.iloc[-21]) - 1) * 100 if len(close) >= 21 else 0

            rows.append({
                "카테고리": category,
                "이름": name,
                "티커": ticker,
                "현재가": round(close.iloc[-1], 2),
                "1일%": round(pct_1d, 2),
                "5일%": round(pct_5d, 2),
                "20일%": round(pct_20d, 2),
                "ATR(14)": round(atr_latest, 2),
                "ATR%": round(atr_pct, 2),
                "RVOL": round(rvol, 2),
                "거래대금($M)": round(dollar_vol_today / 1e6, 1),
                "평균거래대금($M)": round(dollar_vol_avg / 1e6, 1),
                "OBV(M)": round(obv_latest / 1e6, 1),
                "OBV 5일%": round(obv_5d_chg, 1),
            })
        except Exception as e:
            print(f"  ✗ {ticker} 오류: {e}")
            continue

    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════
# 열기 점수 (Heat Score) 계산
# ══════════════════════════════════════════════════════════════

def calc_heat_score(df: pd.DataFrame) -> pd.DataFrame:
    """RVOL, ATR%, 거래대금 비율을 종합한 0~100 점수"""
    if df.empty:
        return df

    df = df.copy()
    # 각 지표 정규화 (0~1)
    for col in ["ATR%", "RVOL"]:
        mn, mx = df[col].min(), df[col].max()
        df[f"{col}_norm"] = (df[col] - mn) / (mx - mn) if mx > mn else 0.5

    vol_ratio = df["거래대금($M)"] / df["평균거래대금($M)"]
    mn, mx = vol_ratio.min(), vol_ratio.max()
    df["vol_ratio_norm"] = (vol_ratio - mn) / (mx - mn) if mx > mn else 0.5

    # 가중 합산: RVOL 40%, ATR% 30%, 거래대금비율 30%
    df["열기점수"] = (
        df["RVOL_norm"] * 40 +
        df["ATR%_norm"] * 30 +
        df["vol_ratio_norm"] * 30
    ).round(1)

    df.drop(columns=["RVOL_norm", "ATR%_norm", "vol_ratio_norm"], inplace=True)
    df.sort_values("열기점수", ascending=False, inplace=True)
    return df


# ══════════════════════════════════════════════════════════════
# Excel 보고서 생성
# ══════════════════════════════════════════════════════════════

# 스타일 상수
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
DATA_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center")

# 열기 등급 색상
HEAT_COLORS = {
    80: PatternFill("solid", fgColor="FF4444"),   # 🔴 극열
    60: PatternFill("solid", fgColor="FF8C00"),   # 🟠 고열
    40: PatternFill("solid", fgColor="FFD700"),   # 🟡 보통
    20: PatternFill("solid", fgColor="90EE90"),   # 🟢 저열
    0:  PatternFill("solid", fgColor="D3D3D3"),   # ⚪ 냉각
}


def get_heat_fill(score):
    for threshold in sorted(HEAT_COLORS.keys(), reverse=True):
        if score >= threshold:
            return HEAT_COLORS[threshold]
    return HEAT_COLORS[0]


def write_data_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str):
    """DataFrame을 포맷팅된 시트로 기록"""
    ws = wb.create_sheet(sheet_name)

    # 표시할 컬럼
    display_cols = [
        "이름", "티커", "현재가", "1일%", "5일%", "20일%",
        "ATR%", "RVOL", "거래대금($M)", "OBV 5일%", "열기점수"
    ]
    cols = [c for c in display_cols if c in df.columns]

    # 헤더
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 데이터
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(cols, 1):
            val = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER

            # 숫자 포맷
            if col_name in ["1일%", "5일%", "20일%", "ATR%", "OBV 5일%"]:
                cell.number_format = "0.00"
                if isinstance(val, (int, float)) and val < 0:
                    cell.font = Font(name="Arial", size=10, color="FF0000")
                elif isinstance(val, (int, float)) and val > 0:
                    cell.font = Font(name="Arial", size=10, color="008000")
            elif col_name == "RVOL":
                cell.number_format = "0.00"
                if isinstance(val, (int, float)) and val >= 2.0:
                    cell.font = Font(name="Arial", size=10, bold=True, color="FF4444")
            elif col_name == "거래대금($M)":
                cell.number_format = "#,##0.0"
            elif col_name == "현재가":
                cell.number_format = "#,##0.00"
            elif col_name == "열기점수":
                cell.number_format = "0.0"
                if isinstance(val, (int, float)):
                    cell.fill = get_heat_fill(val)

    # 열 너비 자동 조정
    for col_idx, col_name in enumerate(cols, 1):
        width = max(len(str(col_name)) * 1.5, 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 필터 설정
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df) + 1}"

    return ws, cols


def add_heat_chart(wb: Workbook, ws, df: pd.DataFrame, cols: list, sheet_name: str):
    """열기점수 막대 차트 추가"""
    if "열기점수" not in cols or df.empty:
        return

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"{sheet_name} — 열기점수 순위"
    chart.y_axis.title = "열기점수 (0~100)"
    chart.x_axis.title = None
    chart.width = 28
    chart.height = 14

    name_col = cols.index("이름") + 1
    score_col = cols.index("열기점수") + 1
    n_rows = len(df)

    data_ref = Reference(ws, min_col=score_col, min_row=1, max_row=n_rows + 1)
    cats_ref = Reference(ws, min_col=name_col, min_row=2, max_row=n_rows + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4

    # 차트 시트에 배치
    chart_ws = wb.create_sheet(f"{sheet_name} 차트")
    chart_ws.add_chart(chart, "A1")


def add_rvol_chart(wb: Workbook, ws, df: pd.DataFrame, cols: list, sheet_name: str):
    """RVOL 막대 차트 (기준선 1.0 강조)"""
    if "RVOL" not in cols or df.empty:
        return

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"{sheet_name} — RVOL (평소 대비 거래량)"
    chart.y_axis.title = "RVOL (1.0 = 평소 수준)"
    chart.width = 28
    chart.height = 14

    name_col = cols.index("이름") + 1
    rvol_col = cols.index("RVOL") + 1
    n_rows = len(df)

    data_ref = Reference(ws, min_col=rvol_col, min_row=1, max_row=n_rows + 1)
    cats_ref = Reference(ws, min_col=name_col, min_row=2, max_row=n_rows + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    chart_ws_name = f"{sheet_name} 차트"
    if chart_ws_name in wb.sheetnames:
        wb[chart_ws_name].add_chart(chart, "A18")
    else:
        chart_ws = wb.create_sheet(chart_ws_name)
        chart_ws.add_chart(chart, "A1")


def add_summary_sheet(wb: Workbook, all_dfs: dict):
    """카테고리별 1위 요약"""
    ws = wb.active
    ws.title = "📊 요약"

    # 제목
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"시장 열기 보고서 — {datetime.now().strftime('%Y-%m-%d %H:%M KST')}"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    ws["A2"].value = "열기점수 = RVOL(40%) + ATR%(30%) + 거래대금비율(30%)  |  🔴≥80 극열  🟠≥60 고열  🟡≥40 보통  🟢≥20 저열  ⚪<20 냉각"
    ws["A2"].font = Font(name="Arial", size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:I3")
    ws["A3"].value = "※ 열기점수는 카테고리 내 상대 순위입니다. 카테고리 간 점수를 직접 비교하지 마세요."
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="999999")
    ws["A3"].alignment = Alignment(horizontal="center")

    # 카테고리별 1위 테이블
    headers = ["카테고리", "최열 종목", "열기점수", "RVOL", "ATR%", "거래대금($M)", "OBV 5일%", "1일%", "5일%"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    row_idx = 6
    for cat_name, df in all_dfs.items():
        if df.empty:
            continue
        top = df.iloc[0]  # 이미 열기점수 내림차순 정렬됨
        vals = [cat_name, top["이름"], top["열기점수"], top["RVOL"],
                top["ATR%"], top["거래대금($M)"], top["OBV 5일%"], top["1일%"], top["5일%"]]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if col_idx == 3 and isinstance(val, (int, float)):
                cell.fill = get_heat_fill(val)
                cell.font = Font(name="Arial", size=10, bold=True)
            if col_idx in [8, 9] and isinstance(val, (int, float)):
                if val < 0:
                    cell.font = Font(name="Arial", size=10, color="FF0000")
                elif val > 0:
                    cell.font = Font(name="Arial", size=10, color="008000")
        row_idx += 1

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # 카테고리별 1위 차트
    n_cats = row_idx - 6
    if n_cats >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "🔥 카테고리별 최열 종목"
        chart.y_axis.title = "열기점수 (카테고리 내)"
        chart.width = 24
        chart.height = 14
        data_ref = Reference(ws, min_col=3, min_row=5, max_row=5 + n_cats)
        cats_ref = Reference(ws, min_col=2, min_row=6, max_row=5 + n_cats)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"A{row_idx + 2}")


def add_history_sheet(wb: Workbook, all_dfs: dict, output_dir: Path):
    """과거 데이터 누적 시트 (JSON 기반)"""
    ws = wb.create_sheet("📈 추이 데이터")
    history_file = output_dir / "heat_history.json"

    # 이전 데이터 로드
    history = []
    if history_file.exists():
        try:
            history = json.loads(history_file.read_text(encoding="utf-8"))
        except Exception:
            history = []

    # 오늘 데이터 추가
    today = datetime.now().strftime("%Y-%m-%d")
    combined = pd.concat(all_dfs.values(), ignore_index=True)
    for _, row in combined.iterrows():
        history.append({
            "date": today,
            "name": row["이름"],
            "category": row["카테고리"],
            "ticker": row["티커"],
            "heat_score": row["열기점수"],
            "rvol": row["RVOL"],
            "atr_pct": row["ATR%"],
            "dollar_vol_m": row["거래대금($M)"],
            "obv_5d_pct": row["OBV 5일%"],
            "pct_1d": row["1일%"],
        })

    # 최근 90일만 유지
    cutoff = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d")
    history = [h for h in history if h["date"] >= cutoff]
    history_file.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")

    # 추이 시트 기록 (피벗: 날짜 × 종목 → 열기점수)
    if not history:
        ws["A1"] = "아직 누적 데이터가 없습니다. 매일 실행하면 여기에 추이가 쌓입니다."
        return

    hist_df = pd.DataFrame(history)
    pivot = hist_df.pivot_table(index="date", columns="name", values="heat_score", aggfunc="last")
    pivot = pivot.sort_index()

    # 헤더
    ws.cell(row=1, column=1, value="날짜").font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    for col_idx, col_name in enumerate(pivot.columns, 2):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    # 데이터
    for row_idx, (date_str, row) in enumerate(pivot.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=date_str).font = DATA_FONT
        for col_idx, col_name in enumerate(pivot.columns, 2):
            val = row.get(col_name)
            if pd.notna(val):
                cell = ws.cell(row=row_idx, column=col_idx, value=round(val, 1))
                cell.font = DATA_FONT
                cell.fill = get_heat_fill(val)
                cell.alignment = CENTER

    # 열 너비
    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(pivot.columns) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # 추이 라인 차트 (데이터가 2일 이상이면)
    if len(pivot) >= 2:
        chart = LineChart()
        chart.title = "열기점수 추이 (최근 90일)"
        chart.y_axis.title = "열기점수"
        chart.width = 35
        chart.height = 16
        chart.style = 10
        n_cols = len(pivot.columns)
        n_rows = len(pivot)
        data_ref = Reference(ws, min_col=2, max_col=n_cols + 1, min_row=1, max_row=n_rows + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"A{n_rows + 4}")


def generate_report(output_dir: Path) -> Path:
    """메인 실행: 데이터 수집 → Excel 보고서 생성"""
    output_dir.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")

    print("=" * 50)
    print(f"📊 시장 열기 보고서 생성 — {date_str}")
    print("=" * 50)

    # 데이터 수집
    categories = {
        "국가별": (COUNTRY_ETFS, "국가", False),
        "미국 섹터": (SECTOR_ETFS, "섹터", False),
        "크립토": (CRYPTO_TICKERS, "크립토", True),
        "원자재": (COMMODITY_ETFS, "원자재", False),
    }

    all_dfs = {}
    for cat_name, (tickers, cat_label, crypto) in categories.items():
        print(f"\n[{cat_label}] {len(tickers)}개 티커 수집 중...")
        df = fetch_and_calc(tickers, cat_label, is_crypto=crypto)
        if not df.empty:
            df = calc_heat_score(df)
            all_dfs[cat_name] = df
            print(f"  → {len(df)}개 완료")
        else:
            print(f"  → 데이터 없음")

    if not all_dfs:
        print("\n✗ 수집된 데이터가 없습니다. 네트워크를 확인하세요.")
        return None

    # Excel 생성
    wb = Workbook()
    add_summary_sheet(wb, all_dfs)

    for sheet_name, df in all_dfs.items():
        ws, cols = write_data_sheet(wb, df, sheet_name)
        add_heat_chart(wb, ws, df, cols, sheet_name)
        add_rvol_chart(wb, ws, df, cols, sheet_name)

    add_history_sheet(wb, all_dfs, output_dir)

    # 저장
    filepath = output_dir / f"market_heat_{date_str}.xlsx"
    wb.save(filepath)
    print(f"\n✓ 보고서 저장: {filepath}")

    # 마크다운 요약도 생성 (GitHub README / 이메일 본문용)
    md_path = output_dir / f"market_heat_{date_str}.md"
    write_markdown_summary(all_dfs, md_path)
    print(f"✓ 마크다운 요약: {md_path}")

    return filepath


def write_markdown_summary(all_dfs: dict, path: Path):
    """이메일 본문 / 빠른 확인용 마크다운"""
    today = datetime.now().strftime("%Y-%m-%d")

    lines = [
        f"# 🔥 시장 열기 보고서 — {today}",
        "",
        "## 카테고리별 최열 종목",
        "",
        "| 카테고리 | 최열 종목 | 열기점수 | RVOL | ATR% | OBV 5일% | 1일% |",
        "|:---:|:---|:---:|:---:|:---:|:---:|:---:|",
    ]
    for cat_name, df in all_dfs.items():
        if df.empty:
            continue
        top = df.iloc[0]
        lines.append(
            f"| {cat_name} | {top['이름']} | "
            f"**{top['열기점수']}** | {top['RVOL']} | {top['ATR%']} | {top['OBV 5일%']} | {top['1일%']}% |"
        )

    lines += ["", "## 카테고리별 상세 순위", ""]
    for cat_name, df in all_dfs.items():
        lines.append(f"### {cat_name}")
        lines.append("")
        lines.append("| 순위 | 이름 | 열기점수 | RVOL | ATR% | 1일% |")
        lines.append("|:---:|:---|:---:|:---:|:---:|:---:|")
        for rank, (_, row) in enumerate(df.iterrows(), 1):
            lines.append(
                f"| {rank} | {row['이름']} | {row['열기점수']} | "
                f"{row['RVOL']} | {row['ATR%']} | {row['1일%']}% |"
            )
        lines.append("")

    lines += [
        "---",
        f"*자동 생성: {datetime.now().strftime('%Y-%m-%d %H:%M')} | 열기점수 = RVOL(40%) + ATR%(30%) + 거래대금비율(30%) — 카테고리 내 상대 순위*",
        "*※ 카테고리 간 점수 직접 비교 불가*",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


# ══════════════════════════════════════════════════════════════
# 엔트리 포인트
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="시장 열기 트래커")
    parser.add_argument("--output", default="./reports", help="보고서 저장 디렉토리")
    args = parser.parse_args()
    generate_report(Path(args.output))
